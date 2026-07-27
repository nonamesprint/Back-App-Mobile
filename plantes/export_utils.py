# plantes/export_utils.py
import csv
import io
from datetime import datetime
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Font, PatternFill, Alignment

def export_to_csv(queryset, fields, headers):
    """Exporte les données en CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(headers)
    
    for obj in queryset:
        row = []
        for field in fields:
            value = getattr(obj, field)
            if callable(value):
                value = value()
            if value is None:
                value = ''
            if isinstance(value, datetime):
                value = value.strftime('%d/%m/%Y %H:%M')
            row.append(str(value))
        writer.writerow(row)
    
    return response


def export_to_excel(queryset, fields, headers, sheet_name='Données'):
    """Exporte les données en Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # En-têtes
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Données
    for row_idx, obj in enumerate(queryset, 2):
        for col_idx, field in enumerate(fields, 1):
            value = getattr(obj, field)
            if callable(value):
                value = value()
            if value is None:
                value = ''
            if isinstance(value, datetime):
                value = value.strftime('%d/%m/%Y %H:%M')
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response


def export_to_pdf(queryset, fields, headers, title="Exportation"):
    """Exporte les données en PDF"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    # Créer le PDF
    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    
    # Style du titre
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        alignment=TA_CENTER,
        spaceAfter=20,
    )
    
    elements = []
    
    # Titre
    elements.append(Paragraph(f"<b>{title}</b>", title_style))
    
    # Date
    date_style = ParagraphStyle(
        'DateStyle',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        spaceAfter=15,
    )
    elements.append(Paragraph(f"Généré le: {timezone.now().strftime('%d/%m/%Y à %H:%M')}", date_style))
    
    # Préparer les données pour le tableau
    data = [headers]
    for obj in queryset:
        row = []
        for field in fields:
            value = getattr(obj, field)
            if callable(value):
                value = value()
            if value is None:
                value = ''
            if isinstance(value, datetime):
                value = value.strftime('%d/%m/%Y %H:%M')
            row.append(str(value))
        data.append(row)
    
    # Créer le tableau
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#447e9b')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 20))
    
    # Pied de page
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        alignment=TA_CENTER,
        textColor=colors.grey,
    )
    elements.append(Paragraph("Jardin Botanique - Exportation automatique", footer_style))
    
    doc.build(elements)
    return response



def export_stats_to_excel(stats_data, title="Statistiques du Jardin Botanique"):
    """Exporte toutes les statistiques en Excel avec structure professionnelle"""
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    
    # Supprimer la feuille par défaut
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # ===== FEUILLE 1 : RÉSUMÉ GÉNÉRAL =====
    ws1 = wb.create_sheet("Résumé Général", 0)
    
    # En-tête
    ws1.merge_cells('A1:B1')
    ws1['A1'] = title
    ws1['A1'].font = Font(size=14, bold=True, color='FFFFFF')
    ws1['A1'].fill = PatternFill(start_color='1B5E20', end_color='1B5E20', fill_type='solid')
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws1['A2'] = f"Généré le: {timezone.now().strftime('%d/%m/%Y à %H:%M')}"
    ws1['A2'].font = Font(size=10, italic=True)
    ws1['A2'].alignment = Alignment(horizontal='center')
    ws1.merge_cells('A2:B2')
    
    # Statistiques générales
    row = 4
    ws1['A4'] = "📊 STATISTIQUES GÉNÉRALES"
    ws1['A4'].font = Font(size=12, bold=True, color='FFFFFF')
    ws1['A4'].fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    ws1.merge_cells(f'A{row}:B{row}')
    row += 1
    
    generales = {
        'Plantes actives': stats_data.get('total_plantes', 0),
        'Circuits actifs': stats_data.get('total_circuits', 0),
        'Articles publiés': stats_data.get('total_articles', 0),
        'Abonnés actifs': stats_data.get('total_abonnes', 0),
        'Réservations totales': stats_data.get('total_reservations', 0),
        'Questions Quiz': stats_data.get('total_questions_quiz', 0),
    }
    for key, value in generales.items():
        if value is not None:
            ws1[f'A{row}'] = key
            ws1[f'B{row}'] = value
            row += 1
    
    # Réservations
    row += 1
    ws1[f'A{row}'] = "📅 RÉSERVATIONS"
    ws1[f'A{row}'].font = Font(size=12, bold=True, color='FFFFFF')
    ws1[f'A{row}'].fill = PatternFill(start_color='E65100', end_color='E65100', fill_type='solid')
    ws1.merge_cells(f'A{row}:B{row}')
    row += 1
    
    reservations_data = {
        'En attente': stats_data.get('reservations_attente', 0),
        'Confirmées': stats_data.get('reservations_confirmees', 0),
        'Annulées': stats_data.get('reservations_annulees', 0),
        'Terminées': stats_data.get('reservations_terminees', 0),
        'Nouvelles (7j)': stats_data.get('nouvelles_reservations', 0),
    }
    for key, value in reservations_data.items():
        ws1[f'A{row}'] = key
        ws1[f'B{row}'] = value
        row += 1
    
    # Réservations par type
    reservations_par_type = stats_data.get('reservations_par_type', {})
    if reservations_par_type:
        ws1[f'A{row}'] = "--- Par type ---"
        ws1[f'A{row}'].font = Font(italic=True)
        row += 1
        for type_name, count in reservations_par_type.items():
            ws1[f'A{row}'] = type_name.upper()
            ws1[f'B{row}'] = count
            row += 1
    
    # Articles
    row += 1
    ws1[f'A{row}'] = "📝 ARTICLES"
    ws1[f'A{row}'].font = Font(size=12, bold=True, color='FFFFFF')
    ws1[f'A{row}'].fill = PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid')
    ws1.merge_cells(f'A{row}:B{row}')
    row += 1
    
    articles_data = {
        'Total publiés': stats_data.get('total_articles', 0),
        'Vues totales': stats_data.get('total_vues_articles', 0),
        'Moyenne vues': stats_data.get('moyenne_vues', 0),
    }
    for key, value in articles_data.items():
        ws1[f'A{row}'] = key
        ws1[f'B{row}'] = value
        row += 1
    
    # Top 5 articles
    articles_pop = stats_data.get('articles_populaires', [])
    if articles_pop:
        ws1[f'A{row}'] = "--- Top 5 Articles ---"
        ws1[f'A{row}'].font = Font(italic=True)
        row += 1
        for idx, article in enumerate(articles_pop, 1):
            ws1[f'A{row}'] = f"#{idx} {article.get('titre', '')}"
            ws1[f'B{row}'] = f"{article.get('vues', 0)} vues"
            row += 1
    
    # Abonnés
    row += 1
    ws1[f'A{row}'] = "👤 ABONNÉS"
    ws1[f'A{row}'].font = Font(size=12, bold=True, color='FFFFFF')
    ws1[f'A{row}'].fill = PatternFill(start_color='6A1B9A', end_color='6A1B9A', fill_type='solid')
    ws1.merge_cells(f'A{row}:B{row}')
    row += 1
    
    abonnes_data = {
        'Actifs': stats_data.get('abonnes_actifs', 0),
        'Inactifs': stats_data.get('abonnes_inactifs', 0),
        'Nouveaux (7j)': stats_data.get('nouveaux_abonnes', 0),
        'Nouveaux (30j)': stats_data.get('nouveaux_abonnes_30j', 0),
    }
    for key, value in abonnes_data.items():
        ws1[f'A{row}'] = key
        ws1[f'B{row}'] = value
        row += 1
    
   
    
    # Ajuster les colonnes
    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 25
    
    # ===== FEUILLE 2 : ÉVOLUTION HEBDOMADAIRE =====
    ws2 = wb.create_sheet("Évolution Hebdomadaire")
    
    ws2['A1'] = "📈 Évolution des réservations (7 derniers jours)"
    ws2['A1'].font = Font(size=14, bold=True)
    ws2.merge_cells('A1:B1')
    
    headers = ['Date', 'Réservations']
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='E65100', end_color='E65100', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    evolution = stats_data.get('reservations_jour', [])
    if evolution:
        for idx, item in enumerate(evolution, 3):
            ws2.cell(row=idx, column=1, value=item.get('date', ''))
            ws2.cell(row=idx, column=2, value=item.get('count', 0))
    
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 20
    
    # ===== FEUILLE 3 : DERNIÈRES RÉSERVATIONS =====
    ws3 = wb.create_sheet("Dernières Réservations")
    
    ws3['A1'] = "🕐 Dernières réservations"
    ws3['A1'].font = Font(size=14, bold=True)
    ws3.merge_cells('A1:F1')
    
    headers = ['Référence', 'Client', 'Email', 'Type', 'Date', 'Statut']
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='447e9b', end_color='447e9b', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    dernieres = stats_data.get('dernieres_reservations', [])
    if dernieres:
        for idx, res in enumerate(dernieres, 3):
            ws3.cell(row=idx, column=1, value=res.get('reference', ''))
            ws3.cell(row=idx, column=2, value=res.get('nom', ''))
            ws3.cell(row=idx, column=3, value=res.get('email', ''))
            ws3.cell(row=idx, column=4, value=res.get('type_visite', ''))
            ws3.cell(row=idx, column=5, value=res.get('date_visite', ''))
            ws3.cell(row=idx, column=6, value=res.get('statut', ''))
    
    ws3.column_dimensions['A'].width = 15
    ws3.column_dimensions['B'].width = 25
    ws3.column_dimensions['C'].width = 30
    ws3.column_dimensions['D'].width = 20
    ws3.column_dimensions['E'].width = 15
    ws3.column_dimensions['F'].width = 15
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="statistiques_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response
# plantes/export_utils.py - Ajouter à la fin

def export_stats_to_csv(stats_data, title="Statistiques du Jardin Botanique"):
    """Exporte toutes les statistiques en CSV structuré"""
    import csv
    from django.http import HttpResponse
    from django.utils import timezone
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="statistiques_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    
    writer.writerow([title])
    writer.writerow([f"Généré le: {timezone.now().strftime('%d/%m/%Y à %H:%M')}"])
    writer.writerow([])
    
    # ===== STATISTIQUES GÉNÉRALES =====
    writer.writerow(["=== STATISTIQUES GÉNÉRALES ==="])
    generales = {
        'Plantes actives': stats_data.get('total_plantes', 0),
        'Circuits actifs': stats_data.get('total_circuits', 0),
        'Articles publiés': stats_data.get('total_articles', 0),
        'Abonnés actifs': stats_data.get('total_abonnes', 0),
        'Réservations totales': stats_data.get('total_reservations', 0),
        'Questions Quiz': stats_data.get('total_questions_quiz', 0),
    }
    writer.writerow(["Métrique", "Valeur"])
    for key, value in generales.items():
        writer.writerow([key, value])
    writer.writerow([])
    
    # ===== RÉSERVATIONS =====
    writer.writerow(["=== RÉSERVATIONS ==="])
    reservations_data = {
        'En attente': stats_data.get('reservations_attente', 0),
        'Confirmées': stats_data.get('reservations_confirmees', 0),
        'Annulées': stats_data.get('reservations_annulees', 0),
        'Terminées': stats_data.get('reservations_terminees', 0),
        'Nouvelles (7j)': stats_data.get('nouvelles_reservations', 0),
    }
    writer.writerow(["Métrique", "Valeur"])
    for key, value in reservations_data.items():
        writer.writerow([key, value])
    
    # Réservations par type
    reservations_par_type = stats_data.get('reservations_par_type', {})
    if reservations_par_type:
        writer.writerow(["--- Par type ---"])
        for type_name, count in reservations_par_type.items():
            writer.writerow([type_name.upper(), count])
    writer.writerow([])
    
    # ===== ARTICLES =====
    writer.writerow(["=== ARTICLES ==="])
    articles_data = {
        'Total publiés': stats_data.get('total_articles', 0),
        'Vues totales': stats_data.get('total_vues_articles', 0),
        'Moyenne vues': stats_data.get('moyenne_vues', 0),
    }
    writer.writerow(["Métrique", "Valeur"])
    for key, value in articles_data.items():
        writer.writerow([key, value])
    
    # Top 5 articles
    articles_pop = stats_data.get('articles_populaires', [])
    if articles_pop:
        writer.writerow(["--- Top 5 Articles ---"])
        writer.writerow(["#", "Titre", "Vues"])
        for idx, article in enumerate(articles_pop, 1):
            writer.writerow([idx, article.get('titre', ''), article.get('vues', 0)])
    writer.writerow([])
    
    # ===== ABONNÉS =====
    writer.writerow(["=== ABONNÉS ==="])
    abonnes_data = {
        'Actifs': stats_data.get('abonnes_actifs', 0),
        'Inactifs': stats_data.get('abonnes_inactifs', 0),
        'Nouveaux (7j)': stats_data.get('nouveaux_abonnes', 0),
        'Nouveaux (30j)': stats_data.get('nouveaux_abonnes_30j', 0),
    }
    writer.writerow(["Métrique", "Valeur"])
    for key, value in abonnes_data.items():
        writer.writerow([key, value])
    writer.writerow([])
    
    # ===== ÉVOLUTION =====
    writer.writerow(["=== ÉVOLUTION HEBDOMADAIRE ==="])
    writer.writerow(["Date", "Réservations"])
    evolution = stats_data.get('reservations_jour', [])
    for item in evolution:
        writer.writerow([item.get('date', ''), item.get('count', 0)])
    
    return response


def export_stats_to_pdf(stats_data, title="Statistiques du Jardin Botanique"):
    """Exporte toutes les statistiques en PDF structuré"""
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from django.http import HttpResponse
    from django.utils import timezone
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="statistiques_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        alignment=TA_CENTER,
        spaceAfter=8,
        textColor=colors.HexColor('#1B5E20'),
    )
    
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        spaceAfter=20,
        textColor=colors.grey,
    )
    
    section_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontSize=13,
        bold=True,
        spaceAfter=10,
        textColor=colors.HexColor('#2E7D32'),
    )
    
    elements = []
    
    elements.append(Paragraph(f"<b>{title}</b>", title_style))
    elements.append(Paragraph(f"Généré le: {timezone.now().strftime('%d/%m/%Y à %H:%M')}", subtitle_style))
    
    # ===== STATISTIQUES GÉNÉRALES =====
    elements.append(Paragraph("📊 Statistiques Générales", section_style))
    generales = {
        'Plantes actives': stats_data.get('total_plantes', 0),
        'Circuits actifs': stats_data.get('total_circuits', 0),
        'Articles publiés': stats_data.get('total_articles', 0),
        'Abonnés actifs': stats_data.get('total_abonnes', 0),
        'Réservations totales': stats_data.get('total_reservations', 0),
        'Questions Quiz': stats_data.get('total_questions_quiz', 0),
    }
    data = [["Métrique", "Valeur"]]
    for key, value in generales.items():
        data.append([key, str(value)])
    
    table = Table(data, colWidths=[300, 200])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#447e9b')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 15))
    
    # ===== RÉSERVATIONS =====
    elements.append(Paragraph("📅 Réservations", section_style))
    reservations_data = {
        'En attente': stats_data.get('reservations_attente', 0),
        'Confirmées': stats_data.get('reservations_confirmees', 0),
        'Annulées': stats_data.get('reservations_annulees', 0),
        'Terminées': stats_data.get('reservations_terminees', 0),
        'Nouvelles (7j)': stats_data.get('nouvelles_reservations', 0),
    }
    data = [["Métrique", "Valeur"]]
    for key, value in reservations_data.items():
        data.append([key, str(value)])
    
    table = Table(data, colWidths=[300, 200])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E65100')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 15))
    
    # ===== ARTICLES =====
    elements.append(Paragraph("📝 Articles", section_style))
    articles_data = {
        'Total publiés': stats_data.get('total_articles', 0),
        'Vues totales': stats_data.get('total_vues_articles', 0),
        'Moyenne vues': stats_data.get('moyenne_vues', 0),
    }
    data = [["Métrique", "Valeur"]]
    for key, value in articles_data.items():
        data.append([key, str(value)])
    
    table = Table(data, colWidths=[300, 200])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0D47A1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 15))
    
    # ===== ABONNÉS =====
    elements.append(Paragraph("👤 Abonnés", section_style))
    abonnes_data = {
        'Actifs': stats_data.get('abonnes_actifs', 0),
        'Inactifs': stats_data.get('abonnes_inactifs', 0),
        'Nouveaux (7j)': stats_data.get('nouveaux_abonnes', 0),
        'Nouveaux (30j)': stats_data.get('nouveaux_abonnes_30j', 0),
    }
    data = [["Métrique", "Valeur"]]
    for key, value in abonnes_data.items():
        data.append([key, str(value)])
    
    table = Table(data, colWidths=[300, 200])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6A1B9A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 15))
    
    # ===== ÉVOLUTION =====
    elements.append(Paragraph("📈 Évolution Hebdomadaire", section_style))
    evolution = stats_data.get('reservations_jour', [])
    if evolution:
        data = [["Date", "Réservations"]]
        for item in evolution:
            data.append([item.get('date', ''), str(item.get('count', 0))])
        
        table = Table(data, colWidths=[200, 150])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E65100')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(table)
    
    elements.append(Spacer(1, 20))
    elements.append(Paragraph("Jardin Botanique - Exportation automatique", 
                              ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER, textColor=colors.grey)))
    
    doc.build(elements)
    return response